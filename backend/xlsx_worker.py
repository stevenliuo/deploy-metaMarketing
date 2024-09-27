import io

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


def get_workbook_sheet_names(file_bytes: bytes) -> list[str]:
    """ Get workbook sheet names """

    file_stream = io.BytesIO(file_bytes)
    workbook = load_workbook(file_stream, read_only=True)

    sheet_names = workbook.sheetnames

    return sheet_names


def get_workbook_sheet_data(file_bytes: bytes, sheet_name: str) -> dict:
    """ Get workbook sheet data """

    file_stream = io.BytesIO(file_bytes)
    workbook = load_workbook(file_stream, read_only=True, data_only=True)

    sheet = workbook[sheet_name]

    result = {}

    # Iterate over all rows and columns in the sheet
    for index, row in enumerate(sheet.iter_rows(), 1):
        for cell_index, cell in enumerate(row, 1):
            if cell.value is None:
                continue
            result.update({f"{get_column_letter(cell_index)}{index}": cell.value})

    return result
